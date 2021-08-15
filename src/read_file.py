import openpyxl

source_path = "../model_file/work_template.xlsx"
own_company_details = []
client_company_details = []
name_of_work = []
purchase_items_data = {}
conditions = []
greeting = ""
sign = []

wb_obj = openpyxl.load_workbook(source_path)
sheet_obj = wb_obj["data_sheet"]

# read own_company_details
for i in range(2, 9):
    cell_obj = sheet_obj.cell(row=i, column=6)
    own_company_details.append(cell_obj.value)
# print(own_company_details)

# read client_company_details
for i in range(10, 14):
    cell_obj = sheet_obj.cell(row=i, column=6)
    client_company_details.append(cell_obj.value)
# print(client_company_details)

# read name_of_work
name_of_work = [sheet_obj.cell(row=15, column=6).value,
                sheet_obj.cell(row=16, column=6).value]
# print(name_of_work)

# read purchase_items_data
m_row = sheet_obj.max_row
# print(m_row)
name = []
qnt = []
rate = []
for i in range(31, m_row+1):
    name.append(sheet_obj.cell(row=i, column=2).value)
    qnt.append(sheet_obj.cell(row=i, column=3).value)
    rate.append(sheet_obj.cell(row=i, column=4).value)
purchase_items_data["name"] = name
purchase_items_data["qnt"] = qnt
purchase_items_data["rate"] = rate
# print(purchase_items_data)

# read footer
conditions = [sheet_obj.cell(row=20, column=1).value, sheet_obj.cell(
    row=21, column=1).value, sheet_obj.cell(row=22, column=1).value]

greeting = sheet_obj.cell(row=23, column=1).value

sign = [sheet_obj.cell(row=21, column=5).value, sheet_obj.cell(
    row=22, column=5).value]

wb_obj.close()
