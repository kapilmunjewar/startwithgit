from read_file import *
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font


def createTable(tablesize):
    main_sheet_obj = wb_obj["main_sheet"]
    # print(tablesize)
    if(tablesize > 11):
        main_sheet_obj.insert_rows(29, tablesize-11)
        # print(tablesize-11)
        sub = "E"+str(18+tablesize-1)
        # print(sub)
        main_sheet_obj.tables['Invoice3'].ref = "A17:"+sub
        # set Total row size
        main_sheet_obj.row_dimensions[tablesize+18].height = 30
    a = 0
    for e in range(18, 18+tablesize):
        main_sheet_obj.row_dimensions[e].height = 30
        main_sheet_obj.cell(
            row=e, column=2).value = purchase_items_data["name"][a]
        main_sheet_obj.cell(
            row=e, column=3).value = purchase_items_data["qnt"][a]
        main_sheet_obj.cell(
            row=e, column=4).value = purchase_items_data["rate"][a]

        main_sheet_obj.cell(e, 1).alignment = Alignment(
            horizontal='center', vertical='center')
        main_sheet_obj.cell(e, 3).alignment = Alignment(
            horizontal='center', vertical='center')
        main_sheet_obj.cell(e, 4).alignment = Alignment(
            horizontal='right', vertical='center')
        main_sheet_obj.cell(e, 5).alignment = Alignment(
            horizontal='right', vertical='center')

        main_sheet_obj.cell(e, 5).number_format = "0.00"
        main_sheet_obj.cell(e, 4).number_format = "0.00"
        a = a+1
        if main_sheet_obj[e][3].value != None:
            main_sheet_obj[e][4].value = float(
                main_sheet_obj[e][3].value)*(main_sheet_obj[e][2].value)
            main_sheet_obj[e][0].value = e - 17
        # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        #                     showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        # main_sheet_obj.tables['Invoice3'].tableStyleInfo = style

    # Add conditions
    if tablesize > 11:
        rwno = tablesize+21
    else:
        rwno = 32
    for e in range(0, 3):
        main_sheet_obj.cell(
            row=(rwno+e), column=1).value = conditions[e]
        main_sheet_obj.merge_cells(
            start_row=rwno+e, start_column=1, end_row=rwno+e, end_column=3)
        main_sheet_obj.cell(rwno+e, 1).font = Font(
            size=11, underline='none', bold=False, italic=False)
        # print(e, conditions[e])

    # Adding sign
    for e in range(0, 2):
        main_sheet_obj.cell(
            row=(rwno+e+1), column=5).value = sign[e]
        main_sheet_obj.cell(rwno+e+1, 5).font = Font(
            size=9 if e == 0 else 12, underline='none', bold=False, italic=False)
        main_sheet_obj.cell(rwno+e+1, 5).alignment = Alignment(
            horizontal='right', vertical='center')
        # print(e, sign[e])

    # Add Greeting
    main_sheet_obj.cell(
        row=(rwno+3), column=1).value = greeting
    main_sheet_obj.merge_cells(
        start_row=rwno+3, start_column=1, end_row=rwno+3, end_column=5)
    main_sheet_obj.cell(rwno+3, 1).font = Font(
        size=11, underline='none', bold=True, italic=False)
    main_sheet_obj.row_dimensions[rwno+3].height = 30
    main_sheet_obj.cell(rwno+3, 1).alignment = Alignment(
        horizontal='center', vertical='center')

    wb_obj.save('../model_file/result.xlsx')
